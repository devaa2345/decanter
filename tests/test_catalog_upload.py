"""
Unit tests for the catalog retrain pipeline (app/catalog_upload.py).

Covers only the pure parsing/diff/keyword-generation logic — nothing here
touches Supabase (create_pending_version, publish_version, etc. all require
a configured client and are exercised manually against a real project, see
SUPABASE_SETUP.md).
"""

from unittest.mock import MagicMock, patch

import pytest

from app.catalog_upload import (
    CatalogParseError,
    ParsedRow,
    _corpus_stopwords,
    _find_column_map,
    _make_unique_id,
    _parse_rows,
    _read_rows,
    CatalogRemovalWarning,
    parse_workbook,
    publish_version,
    _slugify,
    build_catalog_from_rows,
    generate_keywords,
    parse_full_bottle_cell,
)


class TestReadRows:
    """The real sheet has promo/shipping text above the actual table."""

    def test_finds_header_row_past_promo_text(self):
        csv_text = (
            "3% DISCOUNT:,,,\n"
            "On orders from 1200-2000,,,\n"
            "Brand,Fragrance Name,Clone Of,3ml\n"
            "Afnan,9PM Rebel,Creed Aventus Absolu,150\n"
        )
        headers, rows = _read_rows("sheet.csv", csv_text.encode("utf-8"))
        assert headers == ["Brand", "Fragrance Name", "Clone Of", "3ml"]
        assert rows == [["Afnan", "9PM Rebel", "Creed Aventus Absolu", "150"]]

    def test_unsupported_extension_raises(self):
        with pytest.raises(CatalogParseError):
            _read_rows("sheet.pdf", b"whatever")

    def test_no_recognizable_header_raises(self):
        with pytest.raises(CatalogParseError):
            _read_rows("sheet.csv", b"a,b,c\n1,2,3\n")


class TestFindColumnMap:
    def test_standard_headers(self):
        col_map = _find_column_map(["Brand", "Fragrance Name", "Clone Of", "3ml", "10ml", "BNIB"])
        assert col_map["brand"] == 0
        assert col_map["name"] == 1
        assert col_map["clone_of"] == 2
        assert col_map["sizes"] == {"3ml": 3, "10ml": 4}
        assert col_map["full_bottle"] == 5

    def test_synonym_headers_recognized(self):
        col_map = _find_column_map(["Brand", "Perfume Name", "Inspired By", "5ml", "Full Bottle"])
        assert col_map["name"] == 1
        assert col_map["clone_of"] == 2
        assert col_map["full_bottle"] == 4

    def test_case_and_whitespace_insensitive(self):
        col_map = _find_column_map(["brand", " FRAGRANCE NAME ", "10ML"])
        assert "brand" in col_map
        assert "name" in col_map
        assert "10ml" in col_map["sizes"]

    def test_missing_name_column_raises(self):
        with pytest.raises(CatalogParseError):
            _find_column_map(["Brand", "3ml"])

    def test_missing_price_columns_raises(self):
        with pytest.raises(CatalogParseError):
            _find_column_map(["Brand", "Fragrance Name"])


class TestParseFullBottleCell:
    """
    Full-bottle prices come from a free-text cell since the size varies per
    product — this must never guess: anything unclear is a warning.
    """

    def test_empty_cell(self):
        assert parse_full_bottle_cell("") == ({}, [])
        assert parse_full_bottle_cell(None) == ({}, [])

    def test_clean_value(self):
        prices, warnings = parse_full_bottle_cell("100ml - 2800")
        assert prices == {"100ml_full": 2800}
        assert warnings == []

    def test_rupee_symbol_and_thousands_comma(self):
        prices, warnings = parse_full_bottle_cell("100ml ₹2,800")
        assert prices == {"100ml_full": 2800}

    def test_multiple_sizes_in_one_cell(self):
        prices, warnings = parse_full_bottle_cell("50ml/1600, 100ml/2800")
        assert prices == {"50ml_full": 1600, "100ml_full": 2800}

    def test_bare_number_is_a_warning_not_a_guess(self):
        """A price with no size attached must never be silently assigned a size."""
        prices, warnings = parse_full_bottle_cell("2800")
        assert prices == {}
        assert len(warnings) == 1


class TestParseRows:
    def _col_map(self):
        return {"brand": 0, "name": 1, "clone_of": 2, "sizes": {"3ml": 3, "10ml": 4}, "full_bottle": 5}

    def test_basic_row(self):
        rows = [["Afnan", "9PM Rebel", "Creed Aventus Absolu", "150", "390", "100ml - 2800"]]
        parsed, warnings = _parse_rows(rows, self._col_map())
        assert len(parsed) == 1
        assert parsed[0].brand == "Afnan"
        assert parsed[0].clone_of == "Creed Aventus Absolu"
        assert parsed[0].prices == {"3ml": 150, "10ml": 390, "100ml_full": 2800}
        assert warnings == []

    def test_fully_blank_row_skipped_silently(self):
        rows = [["", "", "", "", "", ""]]
        parsed, warnings = _parse_rows(rows, self._col_map())
        assert parsed == []
        assert warnings == []

    def test_brand_only_row_warns_and_skips(self):
        """A stray category-header-style row (e.g. 'Middle Eastern Perfumes') shouldn't become a fake product."""
        rows = [["Middle Eastern Perfumes", "", "", "", "", ""]]
        parsed, warnings = _parse_rows(rows, self._col_map())
        assert parsed == []
        assert len(warnings) == 1

    def test_no_valid_prices_warns_and_skips(self):
        rows = [["Dior", "Sauvage", "", "", "", ""]]
        parsed, warnings = _parse_rows(rows, self._col_map())
        assert parsed == []
        assert "no valid prices" in warnings[0]

    def test_bad_price_warns_but_keeps_other_sizes(self):
        rows = [["Afnan", "Test", "", "abc", "390", ""]]
        parsed, warnings = _parse_rows(rows, self._col_map())
        assert len(parsed) == 1
        assert parsed[0].prices == {"10ml": 390}
        assert any("3ml" in w for w in warnings)

    def test_price_with_commas_and_rupee_symbol(self):
        rows = [["Afnan", "Test", "", "1,250", "₹2,000", ""]]
        parsed, _ = _parse_rows(rows, self._col_map())
        assert parsed[0].prices == {"3ml": 1250, "10ml": 2000}


class TestGenerateKeywords:
    def test_basic_coverage(self):
        kws = generate_keywords("Afnan", "9PM Rebel", "Creed Aventus Absolu", set())
        assert "9pm rebel" in kws
        assert "afnan 9pm rebel" in kws
        assert "creed aventus absolu" in kws
        assert "rebel" in kws

    def test_short_tokens_excluded(self):
        kws = generate_keywords("Al Majed Oud", "Al Majed Oud NOIR", None, set())
        assert "al" not in kws  # below the 3-char floor
        assert "al majed" in kws
        assert "majed" in kws

    def test_generic_stopword_excluded_as_standalone(self):
        kws = generate_keywords("Test Brand", "Test Eau De Parfum", None, set())
        assert "eau" not in kws
        assert "de" not in kws

    def test_corpus_stopword_excluded_standalone_but_bigram_survives(self):
        """A word too common across the catalog (e.g. 'oud') would false-positive-match
        almost everything as a standalone keyword — but 2-word phrases stay specific."""
        kws = generate_keywords("Test Brand", "Test Oud Noir", None, {"oud", "noir"})
        assert "oud" not in kws
        assert "noir" not in kws
        assert "oud noir" in kws

    def test_no_clone_of_does_not_error(self):
        kws = generate_keywords("Brand", "Name", None, set())
        assert isinstance(kws, list) and kws


class TestCorpusStopwords:
    def test_frequent_word_flagged(self):
        rows = [
            ParsedRow(brand="Brand", name=f"Oud Variant {i}", clone_of=None, prices={"3ml": 100})
            for i in range(20)
        ]
        stop = _corpus_stopwords(rows)
        assert "oud" in stop

    def test_rare_word_not_flagged(self):
        rows = [ParsedRow(brand="Brand", name="Unique Rareword", clone_of=None, prices={"3ml": 100})]
        stop = _corpus_stopwords(rows)
        assert "rareword" not in stop


class TestSlugAndIds:
    def test_slugify_matches_existing_catalog_convention(self):
        # Verified against real entries in catalog_data.json
        assert _slugify("Afnan") == "afnan"
        assert _slugify("9PM Rebel") == "9pm_rebel"
        assert _slugify("Al Majed Oud NOIR") == "al_majed_oud_noir"

    def test_make_unique_id_matches_legacy_pattern(self):
        # brand + name concatenated with NO separator between the two slugs —
        # matches "afnan9pm_rebel" already in catalog_data.json
        pid = _make_unique_id("Afnan", "9PM Rebel", used_ids=set())
        assert pid == "afnan9pm_rebel"

    def test_collision_gets_suffixed(self):
        pid = _make_unique_id("Afnan", "9PM Rebel", used_ids={"afnan9pm_rebel"})
        assert pid == "afnan9pm_rebel_2"


class TestBuildCatalogFromRows:
    def test_new_row_is_added(self):
        rows = [ParsedRow(brand="Afnan", name="9PM Rebel", clone_of=None, prices={"3ml": 150})]
        catalog, diff = build_catalog_from_rows(rows, existing={})
        assert diff["added_count"] == 1
        assert diff["updated_count"] == 0
        assert diff["removed_count"] == 0
        assert "afnan9pm_rebel" in catalog

    def test_reupload_with_price_change_updates_same_id(self):
        existing = {
            "afnan9pm_rebel": {
                "display_name": "Afnan 9PM Rebel",
                "prices": {"3ml": 150},
                "keywords": [],
                "clone_of": None,
            }
        }
        rows = [ParsedRow(brand="Afnan", name="9PM Rebel", clone_of=None, prices={"3ml": 160})]
        catalog, diff = build_catalog_from_rows(rows, existing)

        assert diff["added_count"] == 0
        assert diff["updated_count"] == 1
        assert list(catalog.keys()) == ["afnan9pm_rebel"]  # same id reused, not duplicated
        assert catalog["afnan9pm_rebel"]["prices"] == {"3ml": 160}

    def test_reupload_with_no_changes_is_not_flagged_updated(self):
        existing = {
            "afnan9pm_rebel": {
                "display_name": "Afnan 9PM Rebel",
                "prices": {"3ml": 150},
                "keywords": [],
                "clone_of": None,
            }
        }
        rows = [ParsedRow(brand="Afnan", name="9PM Rebel", clone_of=None, prices={"3ml": 150})]
        _, diff = build_catalog_from_rows(rows, existing)
        assert diff["updated_count"] == 0

    def test_row_missing_from_upload_is_flagged_removed(self):
        existing = {
            "afnan9pm_rebel": {
                "display_name": "Afnan 9PM Rebel",
                "prices": {"3ml": 150},
                "keywords": [],
                "clone_of": None,
            }
        }
        _, diff = build_catalog_from_rows([], existing)
        assert diff["removed_count"] == 1
        assert diff["removed"][0]["perfume_id"] == "afnan9pm_rebel"

    def test_new_entries_carry_a_brand_field(self):
        rows = [ParsedRow(brand="Afnan", name="Test", clone_of=None, prices={"3ml": 100})]
        catalog, _ = build_catalog_from_rows(rows, {})
        pid = next(iter(catalog))
        assert catalog[pid]["brand"] == "Afnan"


class TestEndToEnd:
    """Full pipeline: raw CSV bytes -> parsed rows -> candidate catalog + diff."""

    def test_realistic_sheet(self):
        csv_text = (
            "3% DISCOUNT:,,,,,,,,,\n"
            "On orders from 1200-2000,,,,,,,,,\n"
            "Brand,Fragrance Name,Clone Of,3ml,5ml,8ml,10ml,20ml,30ml,BNIB\n"
            "Afnan,9PM Rebel,Creed Aventus Absolu,150,210,310,390,700,950,100ml - 2800\n"
            "Afnan,9PM Night Out,,160,230,340,430,780,1070,\n"
            "Dior,Sauvage,,,,,,,,\n"  # no prices at all -> skipped with a warning
        )
        headers, rows = _read_rows("sheet.csv", csv_text.encode("utf-8"))
        col_map = _find_column_map(headers)
        parsed, warnings = _parse_rows(rows, col_map)

        assert len(parsed) == 2
        assert len(warnings) == 1
        assert "Sauvage" in warnings[0]

        catalog, diff = build_catalog_from_rows(parsed, existing={})
        assert diff["added_count"] == 2
        assert catalog["afnan9pm_rebel"]["prices"]["100ml_full"] == 2800
        assert "creed aventus absolu" in catalog["afnan9pm_rebel"]["keywords"]


class TestAWorkbookIsReadWhole:
    """
    The bug this guards: the uploader read openpyxl's `wb.active` — one
    sheet. For the shop's real six-tab export that is the men's/unisex
    sheet, so uploading their own workbook and publishing it deleted the
    208 products on the women's sheet from the live bot, dropped every
    full-bottle price, and applied none of the men's-wins precedence. No
    error was raised; the diff simply reported "removed: 208".
    """

    HEADER = ["Brand", "Fragrance Name", "Clone Of", "3ml", "5ml", "10ml"]

    @staticmethod
    def _book(sheets: dict) -> bytes:
        import io as _io

        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for title, rows in sheets.items():
            ws = wb.create_sheet(title=title)
            for row in rows:
                ws.append(row)
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_every_product_sheet_is_read_not_just_the_active_one(self):
        content = self._book(
            {
                "Decant Sheet Men-UNISEX": [self.HEADER, ["Afnan", "9PM Rebel", "", 160, 220, 410]],
                "For Women": [self.HEADER, ["Lattafa", "Yara", "", 130, 180, 300]],
            }
        )
        rows, _warnings, report = parse_workbook("book.xlsx", content)
        names = {f"{r.brand} {r.name}" for r in rows}
        assert names == {"Afnan 9PM Rebel", "Lattafa Yara"}
        assert [s["sheet"] for s in report["sheets_read"]] == [
            "Decant Sheet Men-UNISEX",
            "For Women",
        ]

    def test_the_mens_sheet_wins_a_price_disagreement(self):
        """The shop's own rule — see tests/test_import_precedence.py, which
        pins the same rule for the command-line importer. Both paths write
        the same catalog and must not disagree about which row wins."""
        content = self._book(
            {
                "For Women": [self.HEADER, ["Lattafa", "Khamrah", "", 190, 300, 370]],
                "Decant Sheet Men-UNISEX": [self.HEADER, ["Lattafa", "Khamrah", "", 160, 230, 420]],
            }
        )
        rows, warnings, report = parse_workbook("book.xlsx", content)
        assert len(rows) == 1
        assert rows[0].prices["3ml"] == 160
        assert report["duplicate_rows_dropped"] == 1
        assert any("Khamrah" in w for w in warnings)

    def test_a_duplicate_is_reported_with_both_price_sets(self):
        """Dropping a row silently is how the sheet's 39 price conflicts
        stayed invisible for so long."""
        content = self._book(
            {
                "Decant Sheet Men-UNISEX": [self.HEADER, ["Lattafa", "Khamrah", "", 160, 230, 420]],
                "For Women": [self.HEADER, ["Lattafa", "Khamrah", "", 190, 300, 370]],
            }
        )
        _rows, _warnings, report = parse_workbook("book.xlsx", content)
        clash = report["duplicate_details"][0]
        assert clash["kept_prices"]["3ml"] == 160
        assert clash["dropped_prices"]["3ml"] == 190
        assert clash["kept_from"] == "Decant Sheet Men-UNISEX"

    def test_testers_are_never_read(self):
        """Stock-on-hand for the owner. A tester price is not what a
        customer pays and must never reach the catalog."""
        content = self._book(
            {
                "Decant Sheet Men-UNISEX": [self.HEADER, ["Lattafa", "Khamrah", "", 160, 230, 420]],
                "Testers ": [self.HEADER, ["Lattafa", "Tester Only", "", 1, 1, 1]],
            }
        )
        rows, _warnings, _report = parse_workbook("book.xlsx", content)
        assert {r.name for r in rows} == {"Khamrah"}

    def test_a_sheet_with_no_product_table_is_skipped_not_fatal(self):
        content = self._book(
            {
                "Decant Sheet Men-UNISEX": [self.HEADER, ["Lattafa", "Khamrah", "", 160, 230, 420]],
                "Customer Reviews": [["Name", "Review"], ["Aditi", "lovely"]],
            }
        )
        rows, _warnings, _report = parse_workbook("book.xlsx", content)
        assert len(rows) == 1

    def test_a_csv_is_still_one_table(self):
        """Google Sheets exports one tab at a time, so a CSV has no other
        sheets to look for and must keep working exactly as before."""
        csv_text = "Brand,Fragrance Name,3ml\nLattafa,Khamrah,160\n"
        rows, _warnings, report = parse_workbook("sheet.csv", csv_text.encode("utf-8"))
        assert len(rows) == 1
        assert report["sheets_read"] == [{"sheet": "", "rows": 1, "skipped": None}]


class TestFullBottlePricesAreMergedNotGuessed:
    """
    The Retail Packs sheet is shaped differently — a Quantity and a Price
    column rather than a column per size — and the uploader never read it at
    all, so every full-bottle price was lost on upload.
    """

    @staticmethod
    def _book_with_retail(decants: list, retail: list) -> bytes:
        import io as _io

        import openpyxl

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet(title="Decant Sheet Men-UNISEX")
        for row in [["Brand", "Fragrance Name", "Clone Of", "3ml", "5ml"], *decants]:
            ws.append(row)
        rs = wb.create_sheet(title="Retail Packs ")
        for row in [["Brand Name", "Perfume Name", "Quantity (ml)", "Price"], *retail]:
            rs.append(row)
        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_a_full_bottle_price_lands_on_its_decant_product(self):
        content = self._book_with_retail(
            [["Afnan", "9PM Rebel", "", 160, 220]],
            [["AFNAN", "9PM REBEL EDP", 100, 2850]],
        )
        rows, _warnings, report = parse_workbook("book.xlsx", content)
        assert len(rows) == 1
        assert rows[0].prices["100ml_full"] == 2850
        assert report["full_bottle_prices_attached"] == 1

    def test_an_abbreviated_brand_still_matches(self):
        """The retail sheet writes "AHMED" where the decant sheet writes
        "Ahmed Al Maghribi". Comparing whole strings scored those far too
        low to match, so 300+ full-bottle prices went unattached."""
        content = self._book_with_retail(
            [["Ahmed Al Maghribi", "Aqua Oud", "", 170, 230]],
            [["AHMED", "AQUA OUD", 100, 2400]],
        )
        rows, _warnings, _report = parse_workbook("book.xlsx", content)
        assert rows[0].prices["100ml_full"] == 2400

    def test_a_row_matching_nothing_is_reported_never_invented(self):
        """Creating a product from an unmatched retail row would quietly
        double a large part of the catalog — most of them are the same
        perfume written differently, not new stock. So it comes back as
        something for the owner to decide on."""
        content = self._book_with_retail(
            [["Afnan", "9PM Rebel", "", 160, 220]],
            [["NASEEM", "AURORA", 100, 1450]],
        )
        rows, _warnings, report = parse_workbook("book.xlsx", content)
        assert {r.name for r in rows} == {"9PM Rebel"}
        assert report["full_bottle_unmatched"] == 1
        orphan = report["full_bottle_unmatched_details"][0]
        assert orphan["display_name"] == "Naseem Aurora"
        assert orphan["prices"] == {"100ml_full": 1450}

    def test_a_blank_quantity_still_produces_a_sellable_price(self):
        content = self._book_with_retail(
            [["Afnan", "9PM Rebel", "", 160, 220]],
            [["NASEEM", "AURORA", None, 1450]],
        )
        _rows, _warnings, report = parse_workbook("book.xlsx", content)
        assert report["full_bottle_unmatched_details"][0]["prices"] == {"ml_full": 1450}


class TestAPublishThatDeletesTheCatalogIsRefused:
    """
    The last line of defence. A parser bug that drops a sheet shows up as a
    large "removed" count, and on the old diff screen that read as a
    statistic rather than a warning — one click from deleting 208 live
    products. See catalog_upload.MAX_SILENT_REMOVALS.
    """

    def _version(self, removed: int) -> dict:
        return {"id": 7, "status": "pending", "removed_count": removed, "storage_path": "v7.json"}

    def test_a_large_removal_is_refused_until_confirmed(self):
        with patch("app.catalog_upload.require_client", return_value=MagicMock()):
            with patch("app.catalog_upload._get_version", return_value=self._version(208)):
                with patch("app.catalog_upload._activate_version") as activate:
                    with pytest.raises(CatalogRemovalWarning) as excinfo:
                        publish_version(7)
        activate.assert_not_called()
        assert excinfo.value.removed == 208
        assert "208" in str(excinfo.value)

    def test_confirming_lets_it_through(self):
        """The owner is allowed to discontinue a whole line — they just have
        to have read the number first."""
        with patch("app.catalog_upload.require_client", return_value=MagicMock()):
            with patch("app.catalog_upload._get_version", return_value=self._version(208)):
                with patch("app.catalog_upload._activate_version", return_value={"ok": True}) as activate:
                    publish_version(7, confirm_removals=True)
        activate.assert_called_once()

    def test_an_ordinary_publish_needs_no_confirmation(self):
        with patch("app.catalog_upload.require_client", return_value=MagicMock()):
            with patch("app.catalog_upload._get_version", return_value=self._version(2)):
                with patch("app.catalog_upload._activate_version", return_value={"ok": True}) as activate:
                    publish_version(7)
        activate.assert_called_once()
