"""
Catalog data audit — finds rows that look like the SAME perfume entered
twice, plus the name-formatting problems customers can see.

This is about the spreadsheet, not the matcher. When one perfume is on the
sheet under two spellings, the bot is not wrong to return both cards — it
has two products and no way to know they are one. But the customer sees two
cards, often at two different prices, for the thing they asked for once.

WHAT IT LOOKS FOR
-----------------
  duplicate rows     the same brand + name twice on the sheet, which the
                     importer collapses (first one wins) — so the second
                     row's prices are silently discarded
  spelling variant   two names differing only by a typo inside one word:
                     "Classico"/"Clasico", "Amethyst"/"Amythyst"
  brand variant      the same brand spelled two ways across many rows:
                     "Arabiyat Presitige" vs "Arabiyat Prestige"
  extra word         one name carries a word the other does not, and it is
                     not a concentration: "Bade'e Al Sublime" vs "Bade'e Al
                     Oud Sublime"
  formatting         doubled spaces, stray whitespace, and the status
                     markers baked into names ("(Out of Stock)") — all of
                     which are printed to the customer verbatim

WHAT IT DELIBERATELY DOES NOT REPORT
------------------------------------
Concentration and flanker siblings. "Dior Sauvage EDT" and "Dior Sauvage
EDP" are one edit apart and are two different perfumes at two different
prices; so are Asad and Asad Elixir. Reporting those would bury the real
findings under a hundred false ones, so any pair whose only difference is a
word in _CONCENTRATIONS is treated as two genuine products.

Usage:
    python scripts/catalog_audit.py
    python scripts/catalog_audit.py --md CATALOG_AUDIT.md    # a report to read
    python scripts/catalog_audit.py --csv audit.csv          # a list to work through
    python scripts/catalog_audit.py --sheet "Sovereign Scents - Decant Sheet.xlsx"
"""

import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from rapidfuzz import fuzz  # noqa: E402
from rapidfuzz.distance import DamerauLevenshtein  # noqa: E402

from app.catalog import PERFUMES  # noqa: E402
from app import name_index  # noqa: E402

# Words that make two otherwise-identical names two different products.
# A pair differing only by one of these is a sibling, not a duplicate.
_CONCENTRATIONS = {
    "edp", "edt", "edc", "edv", "parfum", "parfums", "perfume", "extrait",
    "elixir", "elixer", "intense", "intensely", "absolu", "absolute",
    "cologne", "og", "pp", "eau", "de", "toilette", "essence", "extreme",
    "exclusif", "exclusive", "limited", "edition", "sport", "fresh",
    "noir", "blanc", "black", "blue", "white", "gold", "red", "rose",
    "oud", "musk", "vanilla", "amber", "tobacco", "leather", "aqua",
    "men", "man", "women", "woman", "homme", "femme", "pour", "for",
    "her", "his", "him", "kids", "junior", "night", "day", "ice", "royal",
}

# Two words are "the same word misspelled" at or under this edit distance,
# provided they are long enough for the distance to mean anything.
_MAX_TYPO_EDITS = 2
_MIN_TYPO_LEN = 5

# Whole-name similarity below which a pair is not worth examining at all.
_PAIR_SIM_MIN = 85.0


def _words(name: str) -> list[str]:
    return name_index.tokenize(name)


def _significant_diff(a_words: list[str], b_words: list[str]) -> tuple[str, str] | None:
    """
    Classify how two names differ. Returns (kind, detail), or None when the
    difference is a concentration/flanker word and they are two real
    products.
    """
    a_extra = list(a_words)
    b_extra = list(b_words)
    for word in list(a_extra):
        if word in b_extra:
            a_extra.remove(word)
            b_extra.remove(word)

    if not a_extra and not b_extra:
        return ("identical name", "same words in the same order")

    # One word swapped for another. A typo if the two are near-identical
    # AND neither is a word that distinguishes real products.
    if len(a_extra) == 1 and len(b_extra) == 1:
        x, y = a_extra[0], b_extra[0]
        if x in _CONCENTRATIONS or y in _CONCENTRATIONS:
            return None
        if (
            min(len(x), len(y)) >= _MIN_TYPO_LEN
            and DamerauLevenshtein.distance(x, y) <= _MAX_TYPO_EDITS
        ):
            return ("spelling variant", f"{x!r} vs {y!r}")
        return None

    # One name has words the other lacks, and nothing was swapped.
    if not a_extra or not b_extra:
        extra = a_extra or b_extra
        if all(w in _CONCENTRATIONS for w in extra):
            return None
        return ("extra word", "one row adds " + ", ".join(repr(w) for w in extra))

    return None


def find_pairs() -> list[dict]:
    """Every pair of catalog rows that looks like one perfume entered twice.
    Compared across the WHOLE catalog rather than bucketed by brand — the
    brand is exactly what a typo like "Franck"/"Frank Oliver" lands in."""
    items = [(pid, " ".join(_words(d["display_name"]))) for pid, d in PERFUMES.items()]
    found = []

    for i, (pid_a, text_a) in enumerate(items):
        if not text_a:
            continue
        for pid_b, text_b in items[i + 1:]:
            if not text_b or abs(len(text_a) - len(text_b)) > 12:
                continue
            if fuzz.ratio(text_a, text_b) < _PAIR_SIM_MIN:
                continue
            verdict = _significant_diff(text_a.split(), text_b.split())
            if verdict is None:
                continue
            kind, detail = verdict
            a, b = PERFUMES[pid_a], PERFUMES[pid_b]
            found.append(
                {
                    "kind": kind,
                    "detail": detail,
                    "name_a": a["display_name"],
                    "name_b": b["display_name"],
                    "prices_match": a["prices"] == b["prices"],
                    "prices_a": a["prices"],
                    "prices_b": b["prices"],
                }
            )
    return found


def find_brand_variants() -> list[dict]:
    """A brand spelled two ways across many rows. Reported separately from
    the pairwise findings because one brand typo produces dozens of pairs,
    and the fix is one find-and-replace rather than dozens of decisions."""
    # Two-word brands only. A single leading word throws off far more noise
    # than signal — "afnan"/"anfar", "armaf"/"armani" and "french"/"franck"
    # are each one or two edits apart and are all real, separate brands.
    # Two words agreeing except for one typo is a much stronger claim, and a
    # genuinely misspelled one-word brand still surfaces in the pairwise scan
    # above (which is how "Franck"/"Frank Oliver" was found).
    counts: Counter = Counter()
    for data in PERFUMES.values():
        words = _words(data["display_name"])
        if len(words) > 2:
            counts[" ".join(words[:2])] += 1

    prefixes = [p for p, n in counts.items() if n >= 2]
    out = []
    for i, a in enumerate(prefixes):
        for b in prefixes[i + 1:]:
            if a == b or a.count(" ") != b.count(" "):
                continue
            aw, bw = a.split(), b.split()
            diff = [(x, y) for x, y in zip(aw, bw) if x != y]
            if len(diff) != 1:
                continue
            x, y = diff[0]
            if x in _CONCENTRATIONS or y in _CONCENTRATIONS:
                continue
            if (
                min(len(x), len(y)) >= _MIN_TYPO_LEN
                and DamerauLevenshtein.distance(x, y) <= _MAX_TYPO_EDITS
            ):
                out.append({"a": a, "a_rows": counts[a], "b": b, "b_rows": counts[b]})
    return out


_MENS_SHEET = "Decant Sheet Men-UNISEX"


def find_precedence_losses(sheet_dupes: list[dict]) -> list[dict]:
    """
    Sizes that exist on the sheet but never reach a customer.

    The import rule is that the men's/unisex sheet wins outright, and within
    a sheet the first row wins — one row decides a product's whole price
    list, rather than two rows disagreeing about it. That is the right rule,
    and it has a cost: where the losing row offered a size the winner does
    not, that size is simply gone. Nobody can buy a 30ml of Lattafa Jasoor
    from this bot even though the women's sheet prices one.

    Reported so the fix is obvious — add the missing size to the winning
    row — rather than left as a silent consequence of a list literal's
    order.
    """
    out = []
    for d in sheet_dupes:
        winner = d["copies"][0]
        lost: dict[str, int] = {}
        sources: set[str] = set()
        for c in d["copies"][1:]:
            for size, price in c["prices"].items():
                if size not in winner["prices"]:
                    lost[size] = price
                    sources.add(c["where"])
        if lost:
            out.append(
                {
                    "name": d["name"],
                    "winner": winner["where"],
                    "winner_prices": winner["prices"],
                    "lost": lost,
                    "sources": sorted(sources),
                }
            )
    return out


def find_formatting() -> list[dict]:
    """Name text that will be printed to a customer exactly as written."""
    out = []
    for data in PERFUMES.values():
        name = data["display_name"]
        problems = []
        if "  " in name:
            problems.append("doubled space")
        if name != name.strip():
            problems.append("leading/trailing space")
        lowered = name.lower()
        for marker in ("out of stock", "discontinued", "not available", "sold out"):
            if marker in lowered:
                problems.append(f"status note in the name ({marker!r})")
        if problems:
            out.append({"name": name, "problems": "; ".join(problems)})
    return out


_SIZE_COLUMNS = ["3ml", "5ml", "8ml", "10ml", "20ml", "30ml"]


def _price(cell) -> "int | None":
    text = str(cell if cell is not None else "").strip()
    if not text:
        return None
    try:
        return int(float(text.replace(",", "").replace("\u20b9", "")))
    except ValueError:
        return None


def _classify_copies(copies: list[dict]) -> str:
    """
    How two copies of one product differ. Three outcomes, and the difference
    between the first two is the difference between a decision and a merge:

      disagreement   the same size is priced two ways. Someone has to say
                     which is right; nobody can work it out from the sheet.
      coverage       every size both copies list agrees, one simply lists
                     more of them. Nothing is contradictory — the shorter
                     copy is just missing rows, and the fix is to keep the
                     union rather than pick a side.
      identical      the same product written twice, same prices.
    """
    base = copies[0]["prices"]
    identical = True
    for c in copies[1:]:
        if c["prices"] != base:
            identical = False
        for size, price in c["prices"].items():
            if size in base and base[size] != price:
                return "disagreement"
    return "identical" if identical else "coverage"


def read_sheet_duplicates(path: Path) -> list[dict]:
    """
    Rows written twice on the sheet — the same brand + fragrance name in two
    places. The importer keeps the first and drops the rest (see
    scripts/import_catalog_xlsx.read_decants), so anything the other copy
    said is lost.

    Prices are read here too, and each pair classified (see
    _classify_copies): a real price disagreement needs someone to decide, a
    copy that merely lists more sizes needs merging, and a plain duplicate
    at matching prices costs nothing today. Lumping all three together as
    "conflicts" would put 49 decisions in front of you when only 39 are
    decisions at all.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []
    if not path.exists():
        return []

    wb = load_workbook(path, data_only=True, read_only=True)
    seen: dict[str, list[dict]] = defaultdict(list)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        brand_col = name_col = None
        size_cols: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            labels = {
                str(c).strip().lower().rstrip(): i
                for i, c in enumerate(row)
                if c is not None and str(c).strip()
            }
            if brand_col is None:
                if "brand" in labels and "fragrance name" in labels:
                    brand_col, name_col = labels["brand"], labels["fragrance name"]
                    size_cols = {s: labels[s] for s in _SIZE_COLUMNS if s in labels}
                continue
            brand = str(row[brand_col] or "").strip() if brand_col < len(row) else ""
            name = str(row[name_col] or "").strip() if name_col < len(row) else ""
            if not name:
                continue
            prices = {}
            for size, col in size_cols.items():
                if col < len(row):
                    value = _price(row[col])
                    if value is not None:
                        prices[size] = value
            if not prices:
                continue  # a row with no prices is a heading, not a product
            key = " ".join(_words(brand + " " + name))
            if key:
                seen[key].append(
                    {
                        "label": (brand + " " + name).strip(),
                        "where": sheet_name.strip() + " row " + str(r_idx),
                        "prices": prices,
                    }
                )

    out = []
    for key, copies in sorted(seen.items()):
        if len(copies) < 2:
            continue
        out.append(
            {
                "name": copies[0]["label"],
                "copies": copies,
                "times": len(copies),
                "status": _classify_copies(copies),
            }
        )
    return out


_MD_SIZES = ["3ml", "5ml", "8ml", "10ml", "20ml", "30ml"]


def _price_table(copies: list[dict]) -> list[str]:
    """One markdown table per duplicated product: a column per size, a row
    per copy, with the sizes that disagree marked. Reading two price dicts
    side by side in prose is exactly the thing a table is for."""
    sizes = [s for s in _MD_SIZES if any(s in c["prices"] for c in copies)]
    if not sizes:
        return []

    base = copies[0]["prices"]
    lines = [
        "| copy | " + " | ".join(sizes) + " |",
        "|---|" + "---|" * len(sizes),
    ]
    for i, c in enumerate(copies):
        label = "**in the catalog**" if i == 0 else "discarded"
        cells = []
        for size in sizes:
            mine = c["prices"].get(size)
            text = f"{mine:,}" if mine is not None else "—"
            if i > 0 and mine != base.get(size):
                text = f"**{text}**"
            cells.append(text)
        lines.append(f"| {label} | " + " | ".join(cells) + " |")
    lines.append("")
    return lines


def write_markdown(
    path: Path,
    brands: list[dict],
    by_kind: dict,
    conflicts: list[dict],
    coverage: list[dict],
    harmless: list[dict],
    losses: list[dict],
    formatting: list[dict],
) -> None:
    """The same findings as the console output, ordered by what each costs
    and written to be read rather than scrolled past."""
    spelling = by_kind.get("spelling variant", [])
    extra = by_kind.get("extra word", [])

    out: list[str] = [
        "# Catalog data audit",
        "",
        f"Every place the decant sheet holds the same product under two rows — "
        f"a misspelling, a second sheet, a stray word. Across {len(PERFUMES):,} products.",
        "",
        "The bot is not wrong to show two cards for these: it has two products and no way "
        "to know they are one. Ordered below by what each costs.",
        "",
        "**How duplicates are resolved on import:** the men's/unisex sheet wins over every "
        "other sheet, and within a sheet the row that comes first wins. Testers are never "
        "read — they are stock-on-hand, not a price list. Locked by "
        "`tests/test_import_precedence.py`.",
        "",
        "| | finding | count |",
        "|---|---|---|",
        f"| 🔴 | [Same size, two different prices](#same-size-two-different-prices) | **{len(conflicts)}** |",
        f"| 🟠 | [One copy lists more sizes](#one-copy-lists-more-sizes) | {len(coverage)} |",
        f"| 🟠 | [Sizes dropped by the men's-sheet rule](#sizes-dropped-by-the-mens-sheet-rule) | {len(losses)} |",
        f"| 🟠 | [One perfume, two spellings](#one-perfume-two-spellings) | {len(spelling)} |",
        f"| 🟠 | [A brand spelled two ways](#a-brand-spelled-two-ways) | {len(brands)} |",
        f"| 🔵 | [One row carries a word the other does not](#one-row-carries-a-word-the-other-does-not) | {len(extra)} |",
        f"| ⚪ | [Listed twice at the same prices](#listed-twice-at-the-same-prices) | {len(harmless)} |",
        f"| ⚪ | [Name text the customer sees](#name-text-the-customer-sees) | {len(formatting)} |",
        "",
        "Regenerate with `python scripts/catalog_audit.py --md CATALOG_AUDIT.md`.",
        "",
        "---",
        "",
        "## Same size, two different prices",
        "",
        "The same size priced two ways. The importer keeps the first copy it reads, so the "
        "other price never reaches the bot — whichever row happens to come first is what "
        "every customer is quoted. **Nobody can work these out from the sheet; each one "
        "needs someone to say which price is right.** Bold cells are where the two disagree.",
        "",
    ]

    for d in conflicts:
        out.append(f"### {d['name']}")
        out.append("")
        out.append("  ·  ".join(f"`{c['where']}`" for c in d["copies"]))
        out.append("")
        out.extend(_price_table(d["copies"]))

    out += [
        "---",
        "",
        "## One copy lists more sizes",
        "",
        "No contradiction here — every size both copies list agrees on the price. One copy "
        "is simply missing rows, and the sizes only it has are being dropped on import. "
        "**Merge these rather than choosing between them**: keep every size either copy "
        "offers.",
        "",
    ]

    for d in coverage:
        out.append(f"### {d['name']}")
        out.append("")
        out.append("  ·  ".join(f"`{c['where']}`" for c in d["copies"]))
        out.append("")
        out.extend(_price_table(d["copies"]))

    out += [
        "---",
        "",
        "## One perfume, two spellings",
        "",
        "Two rows whose names differ only by a typo inside one word.",
        "",
        "| row A | row B | differs | prices |",
        "|---|---|---|---|",
    ]
    for p in spelling:
        flag = "same" if p["prices_match"] else "**different**"
        out.append(f"| {p['name_a']} | {p['name_b']} | {p['detail']} | {flag} |")

    out += [
        "",
        "Price detail for the pairs that disagree:",
        "",
    ]
    for p in spelling:
        if p["prices_match"]:
            continue
        out.append(f"- **{p['name_a']}** — {p['prices_a']}")
        out.append(f"- **{p['name_b']}** — {p['prices_b']}")
        out.append("")

    out += [
        "---",
        "",
        "## A brand spelled two ways",
        "",
        "Not a per-product decision — pick the correct spelling and replace it down the "
        "whole column.",
        "",
        "| spelling | rows | spelling | rows |",
        "|---|---|---|---|",
    ]
    for b in brands:
        out.append(f"| {b['a'].title()} | {b['a_rows']} | {b['b'].title()} | {b['b_rows']} |")

    out += [
        "",
        "---",
        "",
        "## One row carries a word the other does not",
        "",
        "The grey zone. Most are probably real flankers — Ana Abiyedh and Ana Abiyedh Rouge "
        "are two perfumes. A few look like one product written twice. Worth a glance, not a "
        "sweep.",
        "",
        "| row A | row B | difference | prices |",
        "|---|---|---|---|",
    ]
    for p in extra:
        flag = "same" if p["prices_match"] else "**different**"
        out.append(f"| {p['name_a']} | {p['name_b']} | {p['detail']} | {flag} |")

    out += [
        "",
        "---",
        "",
        "## Listed twice at the same prices",
        "",
        "Nothing is being lost today — the same product simply appears on two sheets with "
        "matching prices. Worth tidying only so the two cannot drift apart later, which is "
        "how the price conflicts at the top of this file happened.",
        "",
        "| product | where |",
        "|---|---|",
    ]
    for d in harmless:
        where = " · ".join(f"`{c['where']}`" for c in d["copies"])
        out.append(f"| {d['name']} | {where} |")

    out += [
        "",
        "---",
        "",
        "## Name text the customer sees",
        "",
        "Display names are printed to WhatsApp exactly as written, doubled spaces and status "
        "notes included. A customer asking for Tom Ford Noir gets a card headed "
        "`TOM FORD NOIR EDP(DISCONTINUED GEM)`.",
        "",
        "| name | problem |",
        "|---|---|",
    ]
    for f in formatting:
        out.append(f"| `{f['name']}` | {f['problems']} |")

    out += [
        "",
        "---",
        "",
        "## What this deliberately leaves out",
        "",
        "Concentration and flanker siblings. Dior Sauvage EDT and Dior Sauvage EDP are one "
        "edit apart and are two real perfumes at two real prices; so are Asad and Asad "
        "Elixir. Reporting those would bury everything above under a hundred false ones, so "
        "any pair whose only difference is a concentration word is treated as two genuine "
        "products.",
        "",
    ]

    path.write_text("\n".join(out), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--csv", type=Path, help="write every finding to this file")
    parser.add_argument("--md", type=Path, help="write a readable report to this file")
    parser.add_argument(
        "--sheet",
        type=Path,
        default=Path("Sovereign Scents - Decant Sheet.xlsx"),
        help="the source workbook, for the duplicate-row check",
    )
    args = parser.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    name_index.build_index()

    print("=" * 100)
    print(f"  CATALOG AUDIT — {len(PERFUMES)} products")
    print("=" * 100)

    brands = find_brand_variants()
    print(f"\n  BRAND SPELLED TWO WAYS ({len(brands)})")
    print("  " + "-" * 96)
    for b in brands:
        print(f"    {b['a']!r} ({b['a_rows']} rows)   vs   {b['b']!r} ({b['b_rows']} rows)")
    if not brands:
        print("    none")

    pairs = find_pairs()
    by_kind: dict[str, list[dict]] = defaultdict(list)
    for p in pairs:
        by_kind[p["kind"]].append(p)

    for kind in ("identical name", "spelling variant", "extra word"):
        rows = by_kind.get(kind, [])
        print(f"\n  {kind.upper()} ({len(rows)})")
        print("  " + "-" * 96)
        for p in rows:
            flag = "same prices" if p["prices_match"] else "DIFFERENT PRICES"
            print(f"    {p['name_a']!r}")
            print(f"    {p['name_b']!r}")
            print(f"        {p['detail']}   [{flag}]")
            if not p["prices_match"]:
                print(f"        A: {p['prices_a']}")
                print(f"        B: {p['prices_b']}")
        if not rows:
            print("    none")

    sheet_dupes = read_sheet_duplicates(args.sheet)
    conflicts = [d for d in sheet_dupes if d["status"] == "disagreement"]
    coverage = [d for d in sheet_dupes if d["status"] == "coverage"]
    harmless = [d for d in sheet_dupes if d["status"] == "identical"]

    print("\n  SAME SIZE, TWO DIFFERENT PRICES (%d)" % len(conflicts))
    print("  " + "-" * 96)
    print("    The importer keeps the first copy it reads, so the other price never")
    print("    reaches the catalog. Nobody can work these out from the sheet — each")
    print("    one needs someone to say which price is right.")
    for d in conflicts:
        print("\n    %r" % d["name"])
        for i, c in enumerate(d["copies"]):
            mark = "KEPT   " if i == 0 else "dropped"
            print("        %s  %-40s %s" % (mark, c["where"], c["prices"]))
    if not conflicts:
        print("    none")

    print("\n  ONE COPY LISTS MORE SIZES (%d)" % len(coverage))
    print("  " + "-" * 96)
    print("    No contradiction — every size both copies list agrees. One is simply")
    print("    missing rows, and the sizes only it has are being dropped. Merge, do")
    print("    not choose.")
    for d in coverage:
        print("\n    %r" % d["name"])
        for i, c in enumerate(d["copies"]):
            mark = "KEPT   " if i == 0 else "dropped"
            print("        %s  %-40s %s" % (mark, c["where"], c["prices"]))
    if not coverage:
        print("    none")

    print("\n  SAME PRODUCT LISTED TWICE, SAME PRICES (%d)" % len(harmless))
    print("  " + "-" * 96)
    print("    Harmless to the bot — worth tidying so the two cannot drift apart.")
    for d in harmless:
        print("    %r  —  %s" % (d["name"], ", ".join(c["where"] for c in d["copies"])))
    if not harmless:
        print("    none")

    losses = find_precedence_losses(sheet_dupes)
    print("\n  SIZES DROPPED BY THE MEN'S-SHEET RULE (%d)" % len(losses))
    print("  " + "-" * 96)
    print("    The winning row does not offer a size the losing row did, so nobody")
    print("    can buy it. Add the size to the winning row and it is back.")
    for d in losses:
        sizes = ", ".join("%s %s" % (k, v) for k, v in sorted(d["lost"].items()))
        print("    %-46s loses %-28s (from %s)" % (repr(d["name"]), sizes, "; ".join(d["sources"])))
    if not losses:
        print("    none")

    formatting = find_formatting()
    print(f"\n  NAME FORMATTING ({len(formatting)})")
    print("  " + "-" * 96)
    for f in formatting:
        print(f"    {f['name']!r}  —  {f['problems']}")
    if not formatting:
        print("    none")

    print()
    print("  " + "=" * 96)
    print(f"    brand spelled two ways : {len(brands)}")
    for kind in ("identical name", "spelling variant", "extra word"):
        print(f"    {kind:<22} : {len(by_kind.get(kind, []))}")
    print(f"    {'price disagreements':<22} : {len(conflicts)}   <-- someone has to decide")
    print(f"    {'one copy lists more':<22} : {len(coverage)}   <-- merge, do not choose")
    print(f"    {'listed twice, same price':<22} : {len(harmless)}")
    print(f"    {'sizes dropped by rule':<22} : {len(losses)}   <-- add these to the winning row")
    print(f"    {'formatting':<22} : {len(formatting)}")
    print()

    if args.md:
        write_markdown(
            args.md, brands, by_kind, conflicts, coverage, harmless, losses, formatting
        )
        print(f"  Wrote {args.md}\n")

    if args.csv:
        with args.csv.open("w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            w.writerow(["kind", "name_a", "name_b", "detail", "prices_match", "prices_a", "prices_b"])
            for b in brands:
                w.writerow(
                    ["brand variant", b["a"], b["b"],
                     f"{b['a_rows']} rows vs {b['b_rows']} rows", "", "", ""]
                )
            for p in pairs:
                w.writerow(
                    [p["kind"], p["name_a"], p["name_b"], p["detail"],
                     p["prices_match"], p["prices_a"], p["prices_b"]]
                )
            for d in sheet_dupes:
                kind = {
                    "disagreement": "price disagreement",
                    "coverage": "one copy lists more sizes",
                    "identical": "listed twice, same prices",
                }[d["status"]]
                w.writerow(
                    [
                        kind,
                        d["name"],
                        "",
                        " | ".join(c["where"] + ": " + str(c["prices"]) for c in d["copies"]),
                        d["status"] == "identical",
                        d["copies"][0]["prices"],
                        d["copies"][1]["prices"],
                    ]
                )
            for d in losses:
                w.writerow(
                    [
                        "size dropped by precedence",
                        d["name"],
                        "",
                        "winner " + d["winner"] + " has no " + ", ".join(sorted(d["lost"])),
                        False,
                        d["winner_prices"],
                        d["lost"],
                    ]
                )
            for f in formatting:
                w.writerow(["formatting", f["name"], "", f["problems"], "", "", ""])
        print(f"  Wrote {args.csv}\n")

    return 0


if __name__ == "__main__":
    sys.exit(main())
