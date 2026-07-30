"""
Import the catalog from the real Decant Sheet workbook.

This supersedes both scripts/audit_catalog_against_sheets.py and
scripts/rebuild_catalog_from_sheets.py as the way catalog data gets in. Those
read printed PDFs and had to reconstruct the table from word coordinates; the
workbook has actual cells, so a blank size is unambiguously blank. That
distinction is the whole ballgame — reading prices positionally instead of by
column is what left 110 products in the old catalog with every price shifted
one size tier, quoting the 8ml price for a 5ml decant.

SHEETS USED
-----------
  Decant Sheet Men-UNISEX  Brand | Fragrance Name | Clone Of | 3ml..30ml | BNIB
  For Women                same, without BNIB
  New Decant Additions     same shape — products in neither PDF
  Retail Packs             Brand | Perfume Name | Quantity (ML) | Price
                           -> "<n>ml_full" prices, merged onto the decant
                           entry where one exists

"Testers" is deliberately skipped: it is stock-on-hand for the owner, not a
customer price list.

Existing perfume_ids are reused wherever a display name still matches, so the
analytics history in message_events keeps pointing at the same products.

Run:
    python scripts/import_catalog_xlsx.py            # preview, writes nothing
    python scripts/import_catalog_xlsx.py --apply    # write catalog_data.json
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import logging

logging.disable(logging.CRITICAL)

from rapidfuzz import fuzz, process  # noqa: E402

from app.catalog import CATALOG_PATH, PERFUMES  # noqa: E402
from app.catalog_upload import (  # noqa: E402
    ParsedRow,
    _corpus_stopwords,
    _make_unique_id,
    generate_keywords,
)
from app.matcher import normalize_message  # noqa: E402

WORKBOOK = "Sovereign Scents - Decant Sheet.xlsx"

DECANT_SHEETS = ["Decant Sheet Men-UNISEX", "For Women", "New Decant Additions "]
RETAIL_SHEET = "Retail Packs "

# Products listed on more than one sheet with different prices. The first
# sheet in DECANT_SHEETS wins — the men's/unisex sheet is both far larger and
# more recently priced (it carries sizes and increases the women's sheet
# lacks) — and every clash is written out so the owner can check.
CLASHES: list[tuple[str, str, dict, dict]] = []

SIZE_COLUMNS = ["3ml", "5ml", "8ml", "10ml", "20ml", "30ml"]

# How close a retail-pack name has to be to a decant product to be the same
# thing. The retail sheet is typed in capitals with a concentration suffix
# ("AFNAN 9PM REBEL EDP" for "Afnan 9PM Rebel").
_RETAIL_MATCH = 88
_CONCENTRATION = {"edp", "edt", "edc", "parfum", "extrait", "cologne", "spray"}


def _clean(value) -> str:
    return "" if value is None else str(value).strip()


def _price(value) -> int | None:
    if value is None:
        return None
    text = str(value).replace(",", "").replace("₹", "").strip()
    if not text or text.upper() in {"NA", "N/A", "-"}:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _match_key(name: str) -> str:
    words = [w for w in normalize_message(name).split() if w not in _CONCENTRATION]
    return "".join(words)


def read_decants(wb, notes: list[str]) -> dict[str, ParsedRow]:
    """One entry per product, keyed by normalized display name.

    Sheet order is precedence order: a product appearing on more than one
    sheet keeps the first sheet's prices, and the clash is reported rather
    than silently resolved.
    """
    rows: dict[str, ParsedRow] = {}

    for sheet_name in DECANT_SHEETS:
        if sheet_name not in wb.sheetnames:
            notes.append(f"sheet {sheet_name!r} not found — skipped")
            continue

        ws = wb[sheet_name]
        header_row = None
        columns: dict[str, int] = {}

        for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            labels = {_clean(c).lower().rstrip(): i for i, c in enumerate(row) if _clean(c)}
            if "brand" in labels and "fragrance name" in labels:
                header_row = r_idx
                columns = {
                    "brand": labels["brand"],
                    "name": labels["fragrance name"],
                    "clone": labels.get("clone of", -1),
                }
                for size in SIZE_COLUMNS:
                    if size in labels:
                        columns[size] = labels[size]
                break

        if header_row is None:
            notes.append(f"sheet {sheet_name!r}: no header row — skipped")
            continue

        added = clashed = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            def cell(key: str) -> str:
                idx = columns.get(key, -1)
                return _clean(row[idx]) if 0 <= idx < len(row) else ""

            brand, name = cell("brand"), cell("name")
            if not name:
                continue  # section divider ("Middle Eastern Perfumes") or blank

            prices = {}
            for size in SIZE_COLUMNS:
                idx = columns.get(size)
                if idx is not None and idx < len(row):
                    value = _price(row[idx])
                    if value:
                        prices[size] = value
            if not prices:
                continue  # a name with no prices is a heading, not a product

            key = normalize_message(f"{brand} {name}")
            if key in rows:
                if rows[key].prices != prices:
                    clashed += 1
                    CLASHES.append((f"{brand} {name}".strip(), sheet_name, rows[key].prices, prices))
                continue

            rows[key] = ParsedRow(
                brand=brand, name=name, clone_of=cell("clone") or None, prices=prices
            )
            added += 1

        notes.append(f"{sheet_name.strip()}: {added} products" + (f", {clashed} already listed with different prices (kept the first)" if clashed else ""))

    return rows


def merge_retail_packs(wb, rows: dict[str, ParsedRow], notes: list[str]) -> None:
    """Attach full-bottle prices onto the decant product they belong to."""
    if RETAIL_SHEET not in wb.sheetnames:
        notes.append(f"sheet {RETAIL_SHEET!r} not found — no full-bottle prices")
        return

    ws = wb[RETAIL_SHEET]
    header_row = None
    columns: dict[str, int] = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        labels = {_clean(c).lower().rstrip(): i for i, c in enumerate(row) if _clean(c)}
        if "brand name" in labels and "perfume name" in labels:
            header_row = r_idx
            columns = {
                "brand": labels["brand name"],
                "name": labels["perfume name"],
                "qty": labels.get("quantity (ml)", -1),
                "price": labels.get("price", -1),
            }
            break

    if header_row is None:
        notes.append(f"{RETAIL_SHEET.strip()}: no header row — skipped")
        return

    keys = {_match_key(r.brand + " " + r.name): k for k, r in rows.items()}
    candidates = list(keys)
    attached = orphan = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def cell(key: str) -> str:
            idx = columns.get(key, -1)
            return _clean(row[idx]) if 0 <= idx < len(row) else ""

        brand, name = cell("brand"), cell("name")
        qty, price = _price(cell("qty")), _price(cell("price"))
        if not name or not price or not qty:
            continue

        probe = _match_key(f"{brand} {name}")
        target = keys.get(probe)
        if target is None:
            hit = process.extractOne(probe, candidates, scorer=fuzz.ratio)
            target = keys[hit[0]] if hit and hit[1] >= _RETAIL_MATCH else None
        if target is None:
            orphan += 1
            continue

        rows[target].prices[f"{qty}ml_full"] = price
        attached += 1

    notes.append(
        f"{RETAIL_SHEET.strip()}: {attached} full-bottle prices attached, "
        f"{orphan} sold only as a full bottle (no decant entry)"
    )


def build(rows: dict[str, ParsedRow]) -> dict[str, dict]:
    parsed = list(rows.values())
    corpus_stopwords = _corpus_stopwords(parsed)
    existing = {normalize_message(v["display_name"]): pid for pid, v in PERFUMES.items()}

    catalog: dict[str, dict] = {}
    used: set[str] = set()
    for key, row in rows.items():
        pid = existing.get(key)
        if pid is None or pid in used:
            pid = _make_unique_id(row.brand, row.name, used)
        used.add(pid)
        catalog[pid] = {
            "keywords": generate_keywords(row.brand, row.name, row.clone_of, corpus_stopwords),
            "display_name": f"{row.brand} {row.name}".strip(),
            "brand": row.brand or None,
            "prices": row.prices,
            "clone_of": row.clone_of,
        }
    return catalog


def diff(old: dict, new: dict) -> dict:
    old_by = {normalize_message(v["display_name"]): v for v in old.values()}
    new_by = {normalize_message(v["display_name"]): v for v in new.values()}
    return {
        "added": sorted(new_by.keys() - old_by.keys()),
        "removed": sorted(old_by.keys() - new_by.keys()),
        "changed": sorted(
            k for k in old_by.keys() & new_by.keys()
            if old_by[k].get("prices") != new_by[k].get("prices")
        ),
        "old_by": old_by,
        "new_by": new_by,
    }


def _fmt(prices: dict) -> str:
    order = SIZE_COLUMNS
    parts = [f"{s} {prices[s]}" for s in order if s in prices]
    parts += [f"{s} {v}" for s, v in sorted(prices.items()) if "full" in s]
    return "  ".join(parts) or "(none)"


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--apply", action="store_true", help="write app/catalog_data.json")
    parser.add_argument("--show", type=int, default=15)
    args = parser.parse_args()

    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    import openpyxl

    if not os.path.exists(WORKBOOK):
        print(f"Workbook not found: {WORKBOOK}", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    notes: list[str] = []
    rows = read_decants(wb, notes)
    merge_retail_packs(wb, rows, notes)
    wb.close()

    catalog = build(rows)
    changes = diff(PERFUMES, catalog)

    print()
    print("=" * 92)
    print("  CATALOG IMPORT FROM WORKBOOK" + ("" if args.apply else "  (preview — nothing written)"))
    print("=" * 92)
    print()
    for note in notes:
        print(f"  - {note}")

    print()
    print(f"  Current catalog : {len(PERFUMES)} products")
    print(f"  Imported catalog: {len(catalog)} products")
    print()
    print(f"  NEW products    : {len(changes['added'])}")
    print(f"  PRICE CHANGES   : {len(changes['changed'])}")
    print(f"  No longer listed: {len(changes['removed'])}")

    if changes["changed"]:
        print()
        print(f"  --- Price changes (showing {min(len(changes['changed']), args.show)}) ---")
        for key in changes["changed"][: args.show]:
            print(f"    {changes['new_by'][key]['display_name']}")
            print(f"        was: {_fmt(changes['old_by'][key].get('prices', {}))}")
            print(f"        now: {_fmt(changes['new_by'][key].get('prices', {}))}")

    if changes["added"]:
        print()
        print(f"  --- New products (showing {min(len(changes['added']), args.show)}) ---")
        for key in changes["added"][: args.show]:
            v = changes["new_by"][key]
            print(f"    {v['display_name']:<50} {_fmt(v['prices'])}")

    if changes["removed"]:
        print()
        print(f"  --- No longer on any sheet (showing {min(len(changes['removed']), args.show)}) ---")
        for key in changes["removed"][: args.show]:
            v = changes["old_by"][key]
            print(f"    {v['display_name']:<50} {_fmt(v.get('prices', {}))}")

    if CLASHES:
        with open("catalog_sheet_clashes.txt", "w", encoding="utf-8") as f:
            f.write("Products listed on more than one sheet with DIFFERENT prices.\n")
            f.write("The men's/unisex sheet wins (larger and more recently priced).\n\n")
            for name, sheet, kept, ignored in CLASHES:
                f.write(f"{name}\n")
                f.write(f"    kept   : {_fmt(kept)}\n")
                f.write(f"    ignored ({sheet.strip()}): {_fmt(ignored)}\n\n")
        print()
        print(f"  {len(CLASHES)} products are priced differently on two sheets — see catalog_sheet_clashes.txt")

    if args.apply:
        backup = str(CATALOG_PATH) + ".bak"
        shutil.copy(CATALOG_PATH, backup)
        CATALOG_PATH.write_text(
            json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print()
        print(f"  Wrote {CATALOG_PATH}  (previous version saved to {backup})")
    else:
        print()
        print("  Nothing written. Re-run with --apply to adopt this.")
    print()


if __name__ == "__main__":
    main()
