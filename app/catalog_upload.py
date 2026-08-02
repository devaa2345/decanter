"""
Catalog retrain pipeline: turn an uploaded sheet (.xlsx/.csv) into the same
shape as catalog_data.json, diff it against the live catalog, and let the
owner review the diff before it goes live.

There's no ML model here — "retrain" means re-deriving keywords/prices/ids
from a fresh sheet using the same conventions the existing 1,200+ entry
catalog already follows (reverse-engineered from catalog_data.json):

  - perfume_id  = slug(brand) + slug(fragrance_name), concatenated directly
                  (e.g. "Afnan" + "9PM Rebel" -> "afnan" + "9pm_rebel").
  - display_name = f"{brand} {fragrance_name}".strip()
  - keywords include 1/2/3-word windows of brand+name and of clone_of (so a
    customer typing the *original* designer perfume name still matches the
    clone that's inspired by it).

One deliberate improvement over a naive tokenizer: single-word keywords are
filtered by BOTH the static app.matcher.GENERIC_STOPWORDS list (generic
English filler words) and a per-upload *corpus* frequency check — a word
like "oud" or "noir" that shows up across dozens of entries in a Middle
Eastern-perfume-heavy catalog would otherwise become a standalone keyword
that matches almost everything (the same false-positive problem
GENERIC_STOPWORDS already exists to prevent for English filler words).
Multi-word phrases aren't filtered this way since a 2-3 word phrase is
inherently specific enough to be safe.

parse_upload() only ever *produces* a candidate version + diff — nothing
here touches the live catalog directly except _activate_version(), reached
only via publish_version()/rollback_version().
"""

import csv
import io
import json
import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone

from app.db import CATALOG_BUCKET, require_client
from app.matcher import normalize_message
from app.name_index import GENERIC_STOPWORDS

logger = logging.getLogger(__name__)


class CatalogParseError(Exception):
    """Raised for problems with a specific upload/version request (bad file, bad id, wrong state)."""


class CatalogRemovalWarning(Exception):
    """A publish that would delete a large part of the catalog, refused
    pending confirmation. Carries the count so the caller can say how many
    rather than just that it was 'a lot' — see MAX_SILENT_REMOVALS."""

    def __init__(self, message: str, removed: int) -> None:
        super().__init__(message)
        self.removed = removed


# --- Header recognition -----------------------------------------------------

BRAND_HEADERS = {"brand"}
NAME_HEADERS = {"fragrance name", "perfume name", "name", "product name", "fragrance"}
CLONE_HEADERS = {"clone of", "clone", "inspired by", "original", "inspired"}
FULL_BOTTLE_HEADERS = {"bnib", "full bottle", "full", "fullbottle", "bottle", "full size"}
SIZE_HEADER_RE = re.compile(r"^(\d+(?:\.\d+)?)\s*ml$", re.IGNORECASE)


def _find_column_map(headers: list[str]) -> dict:
    """Map recognized column purposes to their column index."""
    col_map: dict = {"sizes": {}}

    for idx, raw_header in enumerate(headers):
        h = (raw_header or "").strip().lower()
        if not h:
            continue
        if h in BRAND_HEADERS:
            col_map["brand"] = idx
        elif h in NAME_HEADERS:
            col_map["name"] = idx
        elif h in CLONE_HEADERS:
            col_map["clone_of"] = idx
        elif h in FULL_BOTTLE_HEADERS:
            col_map["full_bottle"] = idx
        else:
            m = SIZE_HEADER_RE.match(h)
            if m:
                size_num = m.group(1)
                col_map["sizes"][f"{size_num}ml"] = idx

    if "name" not in col_map:
        raise CatalogParseError(
            "Could not find a 'Fragrance Name' column. Found headers: "
            + ", ".join(h for h in headers if h)
        )
    if not col_map["sizes"] and "full_bottle" not in col_map:
        raise CatalogParseError(
            "Could not find any size/price columns (e.g. '3ml', '10ml', 'BNIB'). Found headers: "
            + ", ".join(h for h in headers if h)
        )

    return col_map


# Sheets in the shop's workbook that are not a customer price list, matched
# case-insensitively on the sheet name. "Testers" is stock-on-hand for the
# owner and a tester price is not what a customer pays; the rest hold no
# product rows at all.
SKIP_SHEETS = {"testers", "sync log", "decant bottle pics", "customer reviews", "retail packs"}

# Precedence between sheets, most authoritative first. A product listed on
# more than one sheet keeps the FIRST sheet's prices — the men's/unisex
# sheet is far larger and more recently priced, and one row deciding a
# product's whole price list beats two rows disagreeing about it. Sheets not
# named here keep their workbook order, after these.
SHEET_PRECEDENCE = ["decant sheet men-unisex", "for women", "new decant additions"]

# The retail sheet uses its own headers and its own shape: one row per
# bottle size rather than a row of size columns.
RETAIL_SHEET_NAME = "retail packs"
RETAIL_BRAND_HEADERS = {"brand name"}
RETAIL_NAME_HEADERS = {"perfume name"}
RETAIL_QTY_HEADERS = {"quantity (ml)", "quantity", "qty (ml)", "qty"}
RETAIL_PRICE_HEADERS = {"price", "mrp", "retail price"}


def _sheet_sort_key(name: str) -> tuple[int, int]:
    cleaned = name.strip().lower()
    try:
        return (0, SHEET_PRECEDENCE.index(cleaned))
    except ValueError:
        return (1, 0)


def _find_header_row(all_rows: list[list]) -> int | None:
    """The real sheet has promo/shipping text above the actual table, so the
    header is scanned for rather than assumed to be row 0."""
    for i, row in enumerate(all_rows[:20]):
        lowered = {str(c).strip().lower().rstrip() for c in row if c}
        if lowered & (BRAND_HEADERS | NAME_HEADERS):
            return i
    return None


def _read_sheets(filename: str, content: bytes) -> list[tuple[str, list[str], list[list]]]:
    """
    Every product sheet in the upload, as (sheet name, headers, data rows),
    most authoritative first.

    This reads the WHOLE workbook. It used to read openpyxl's `wb.active` —
    one sheet — which for the shop's real export means the men's/unisex
    sheet and nothing else. Uploading that file and publishing it deleted
    the 208 products on the women's sheet from the live bot, dropped every
    full-bottle price, and applied none of the men's-wins precedence, all
    without an error: the diff simply said "removed: 208" as though that
    were a fact about the sheet rather than a fact about the parser.

    A CSV is still one table — Google Sheets exports one tab at a time — so
    it comes back as a single unnamed sheet.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if filename and "." in filename else ""

    if ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        all_rows = [row for row in csv.reader(io.StringIO(text))]
        header_idx = _find_header_row(all_rows)
        if header_idx is None:
            raise CatalogParseError(
                "Could not find a header row containing 'Brand' or 'Fragrance Name' "
                "in the first 20 rows."
            )
        return [("", all_rows[header_idx], all_rows[header_idx + 1 :])]

    if ext not in ("xlsx", "xlsm"):
        raise CatalogParseError(
            f"Unsupported file type '.{ext}'. Please upload a .xlsx or .csv export of the sheet "
            "(PDF exports can't be parsed reliably — use File > Download in Google Sheets instead)."
        )

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    found: list[tuple[str, list[str], list[list]]] = []

    for sheet_name in sorted(wb.sheetnames, key=_sheet_sort_key):
        if sheet_name.strip().lower() in SKIP_SHEETS:
            continue
        rows = [
            ["" if c is None else str(c) for c in row]
            for row in wb[sheet_name].iter_rows(values_only=True)
        ]
        header_idx = _find_header_row(rows)
        if header_idx is None:
            continue  # not a product sheet — pictures, reviews, a change log
        found.append((sheet_name.strip(), rows[header_idx], rows[header_idx + 1 :]))

    if not found:
        raise CatalogParseError(
            "No sheet in this workbook has a header row containing 'Brand' or "
            "'Fragrance Name'. Sheets found: " + ", ".join(wb.sheetnames)
        )

    return found


def _read_rows(filename: str, content: bytes) -> tuple[list[str], list[list]]:
    """The first product sheet, as (headers, data rows).

    Kept for callers that only ever deal in one table — a CSV export, or a
    single-sheet check. Anything importing a real workbook wants
    _read_sheets, which returns all of them: reading one sheet of six is
    precisely the bug this module was repaired for.
    """
    sheets = _read_sheets(filename, content)
    _name, headers, data_rows = sheets[0]
    return headers, data_rows


def _read_retail_packs(filename: str, content: bytes) -> list[tuple[str, str, int | None, int]]:
    """
    The Retail Packs sheet as (brand, name, quantity_ml, price).

    Kept separate because it is shaped differently — a Quantity and a Price
    column rather than a column per size — and because it is the only place
    43 of the shop's perfumes appear at all. Returns [] when the sheet is
    absent, which is the normal case for a single-tab CSV export.
    """
    ext = filename.lower().rsplit(".", 1)[-1] if filename and "." in filename else ""
    if ext not in ("xlsx", "xlsm"):
        return []

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet_name = next(
        (n for n in wb.sheetnames if n.strip().lower() == RETAIL_SHEET_NAME), None
    )
    if sheet_name is None:
        return []

    out: list[tuple[str, str, int | None, int]] = []
    cols: dict[str, int] = {}
    for row in wb[sheet_name].iter_rows(values_only=True):
        labels = {
            str(c).strip().lower().rstrip(): i
            for i, c in enumerate(row)
            if c is not None and str(c).strip()
        }
        if not cols:
            brand_col = next((labels[h] for h in RETAIL_BRAND_HEADERS if h in labels), None)
            name_col = next((labels[h] for h in RETAIL_NAME_HEADERS if h in labels), None)
            if brand_col is None or name_col is None:
                continue
            cols = {
                "brand": brand_col,
                "name": name_col,
                "qty": next((labels[h] for h in RETAIL_QTY_HEADERS if h in labels), -1),
                "price": next((labels[h] for h in RETAIL_PRICE_HEADERS if h in labels), -1),
            }
            continue

        def cell(key: str) -> str:
            idx = cols.get(key, -1)
            return str(row[idx]).strip() if 0 <= idx < len(row) and row[idx] is not None else ""

        name = cell("name")
        price = _int_or_none(cell("price"))
        if not name or price is None:
            continue
        out.append((cell("brand"), name, _int_or_none(cell("qty")), price))

    return out


def _int_or_none(text: str) -> int | None:
    cleaned = (text or "").replace(",", "").replace("₹", "").strip()
    if not cleaned or cleaned.upper() in {"NA", "N/A", "-"}:
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


# --- Full-bottle (BNIB) free-text price parsing -----------------------------

_SIZE_PRICE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*ml[^\d]{0,15}?(\d[\d,]*)", re.IGNORECASE)


def parse_full_bottle_cell(raw: str) -> tuple[dict[str, int], list[str]]:
    """
    Parse a free-text full-bottle cell (e.g. "100ml - 2800", "50ml/1600,
    100ml/2800") into {"{size}ml_full": price} entries.

    Anything that doesn't match a clear "SIZEml ... PRICE" pattern is left
    out and reported as a warning instead of guessed — this feeds
    customer-facing prices, so silent guesses aren't an option.
    """
    raw = (raw or "").strip()
    if not raw:
        return {}, []

    matches = list(_SIZE_PRICE_RE.finditer(raw))
    if not matches:
        return {}, [f"couldn't read a size+price from full-bottle value {raw!r} — expected e.g. '100ml - 2800'"]

    prices: dict[str, int] = {}
    for m in matches:
        size_str, price_str = m.group(1), m.group(2).replace(",", "")
        size_num = float(size_str)
        size_key = f"{int(size_num)}ml_full" if size_num.is_integer() else f"{size_str}ml_full"
        try:
            prices[size_key] = int(price_str)
        except ValueError:
            continue

    return prices, []


# --- Row parsing --------------------------------------------------------

@dataclass
class ParsedRow:
    brand: str
    name: str
    clone_of: str | None
    prices: dict[str, int] = field(default_factory=dict)


def _parse_rows(data_rows: list[list], col_map: dict) -> tuple[list[ParsedRow], list[str]]:
    parsed: list[ParsedRow] = []
    warnings: list[str] = []

    def cell(row: list, idx: int | None) -> str:
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    for i, row in enumerate(data_rows):
        row_num = i + 1
        brand = cell(row, col_map.get("brand"))
        name = cell(row, col_map.get("name"))

        if not name and not brand:
            continue  # blank / section-divider row — not an error, just skip

        if not name:
            warnings.append(f"row {row_num}: has brand '{brand}' but no fragrance name — skipped")
            continue

        clone_of = cell(row, col_map.get("clone_of")) or None

        prices: dict[str, int] = {}
        for tier, idx in col_map["sizes"].items():
            raw_val = cell(row, idx)
            if not raw_val:
                continue
            cleaned = raw_val.replace(",", "").replace("₹", "").strip()
            try:
                prices[tier] = int(float(cleaned))
            except ValueError:
                warnings.append(f"row {row_num} ({name}): couldn't read {tier} price {raw_val!r} — skipped that size")

        if "full_bottle" in col_map:
            fb_prices, fb_warnings = parse_full_bottle_cell(cell(row, col_map["full_bottle"]))
            prices.update(fb_prices)
            warnings.extend(f"row {row_num} ({name}): {w}" for w in fb_warnings)

        if not prices:
            warnings.append(f"row {row_num} ({name}): no valid prices for any size — entry skipped")
            continue

        parsed.append(ParsedRow(brand=brand, name=name, clone_of=clone_of, prices=prices))

    return parsed, warnings


def _row_key(row: ParsedRow) -> str:
    return normalize_message(f"{row.brand} {row.name}")


def dedupe_first_wins(
    sheets: list[tuple[str, list[ParsedRow]]],
) -> tuple[list[ParsedRow], list[str], list[dict]]:
    """
    One row per product, resolving the ~90 the shop's sheet lists twice.

    The rule, and it is the shop's own: the first sheet wins over every
    later one, and within a sheet the first row wins. One row decides a
    product's whole price list rather than two rows disagreeing about it.

    Nothing is resolved silently. Every dropped copy comes back in
    `clashes` — with both sets of prices, so the owner can see which
    product is being quoted at which price and fix the sheet rather than
    discovering it from a customer.
    """
    kept: dict[str, ParsedRow] = {}
    origin: dict[str, str] = {}
    warnings: list[str] = []
    clashes: list[dict] = []

    for sheet_name, rows in sheets:
        for row in rows:
            key = _row_key(row)
            if not key:
                continue
            if key not in kept:
                kept[key] = row
                origin[key] = sheet_name
                continue
            if kept[key].prices != row.prices:
                clashes.append(
                    {
                        "display_name": f"{kept[key].brand} {kept[key].name}".strip(),
                        "kept_from": origin[key] or "earlier row",
                        "kept_prices": dict(kept[key].prices),
                        "dropped_from": sheet_name or "later row",
                        "dropped_prices": dict(row.prices),
                    }
                )
                warnings.append(
                    f"{kept[key].brand} {kept[key].name}".strip()
                    + f": listed again on {sheet_name or 'a later row'} with different prices"
                    + " — kept the first"
                )

    return list(kept.values()), warnings, clashes


# The retail sheet is typed in block capitals ("LATTAFA PETRA EDP"). Title
# case reads as a name; these words do not survive it and are restored.
_RETAIL_CAPS = {"edp", "edt", "edc", "pp", "og", "ml", "uae", "usa", "uk", "ii", "iii", "iv"}


def _retail_display(brand: str, name: str) -> tuple[str, str]:
    def fix(text: str) -> str:
        return " ".join(
            w.upper() if w.lower().strip(".,") in _RETAIL_CAPS else w.title()
            for w in text.split()
        )

    return fix(brand), fix(name)


# Words the retail sheet adds or drops freely against the decant sheet.
# Removed from both sides before two names are compared, so "AFNAN 9PM
# REBEL EDP" and "Afnan 9PM Rebel" are recognized as one product.
_RETAIL_FILLER = {
    "pour", "homme", "femme", "for", "men", "women", "man", "woman",
    "eau", "de", "by", "the",
}

# Similarity at which two names are the same product. Applied to the
# PRODUCT part only, with the brand checked separately — the retail sheet
# writes "AHMED" where the decant sheet writes "Ahmed Al Maghribi", which
# drags a whole-string comparison down far below anything usable.
_RETAIL_NAME_MIN = 92


def _retail_name_sig(name: str) -> str:
    return "".join(
        w
        for w in normalize_message(name).split()
        if w not in _RETAIL_CONCENTRATION and w not in _RETAIL_FILLER
    )


def _brands_compatible(a: frozenset, b: frozenset) -> bool:
    """One brand written short and the other written in full is the same
    brand. Two brands that merely overlap are not."""
    return not a or not b or a <= b or b <= a


def merge_retail_packs(
    rows: list[ParsedRow], retail: list[tuple[str, str, int | None, int]]
) -> tuple[list[ParsedRow], int, list[dict]]:
    """
    Fold full-bottle prices into the decant product they belong to, and
    report every retail row that matched nothing.

    Report rather than create, deliberately. 680 rows on this sheet are
    full bottles and roughly 370 of them match no decant product by any
    rule I would trust — but most of those are the same perfume written
    slightly differently ("AHMED BIN SHAIKH" for "Ahmed Al Maghribi Bin
    Sheikh"), not new products. Creating them would quietly double a large
    part of the catalog and attach real prices to the wrong bottle, which
    is the same silent damage this whole module is being repaired for.

    So the unmatched rows come back as data. The dashboard shows them, marks
    the ones the bot genuinely cannot answer at all, and lets the owner add
    those deliberately — a decision made by someone who knows the stock,
    rather than a guess made by a fuzzy ratio.
    """
    from rapidfuzz import fuzz, process

    by_key = {_match_key(f"{r.brand} {r.name}"): r for r in rows}
    by_name: dict[str, list[tuple[frozenset, ParsedRow]]] = {}
    for r in rows:
        by_name.setdefault(_retail_name_sig(r.name), []).append(
            (frozenset(normalize_message(r.brand).split()), r)
        )
    full_keys = list(by_key)
    name_keys = list(by_name)

    attached = 0
    unmatched: list[dict] = []

    for brand, name, qty, price in retail:
        probe = _match_key(f"{brand} {name}")
        target = by_key.get(probe)

        if target is None:
            # Brand and product part compared separately, which is what
            # recovers the abbreviated-brand rows.
            rb = frozenset(normalize_message(brand).split())
            rn = _retail_name_sig(name)
            for db, candidate in by_name.get(rn, []):
                if _brands_compatible(rb, db):
                    target = candidate
                    break
            if target is None and name_keys:
                hit = process.extractOne(rn, name_keys, scorer=fuzz.ratio)
                if hit and hit[1] >= _RETAIL_NAME_MIN:
                    for db, candidate in by_name[hit[0]]:
                        if _brands_compatible(rb, db):
                            target = candidate
                            break
            if target is None and full_keys:
                hit = process.extractOne(probe, full_keys, scorer=fuzz.ratio)
                if hit and hit[1] >= RETAIL_MATCH_MIN:
                    target = by_key[hit[0]]

        # A size the sheet left blank still has to be sellable. The key
        # carries no number, which app.formatter renders as "Full bottle"
        # rather than inventing a millilitre figure nobody wrote down.
        size_key = f"{qty}ml_full" if qty else "ml_full"

        if target is not None:
            target.prices[size_key] = price
            attached += 1
            continue

        disp_brand, disp_name = _retail_display(brand, name)
        unmatched.append(
            {
                "brand": disp_brand,
                "name": disp_name,
                "display_name": f"{disp_brand} {disp_name}".strip(),
                "prices": {size_key: price},
            }
        )

    return rows, attached, unmatched


# How close a retail-pack name has to be to a decant product to be the same
# thing. The retail sheet is capitals with a concentration suffix
# ("AFNAN 9PM REBEL EDP" for "Afnan 9PM Rebel").
RETAIL_MATCH_MIN = 88
_RETAIL_CONCENTRATION = {"edp", "edt", "edc", "parfum", "extrait", "cologne", "spray"}


def _match_key(name: str) -> str:
    words = [w for w in normalize_message(name).split() if w not in _RETAIL_CONCENTRATION]
    return "".join(words)


def parse_workbook(
    filename: str, content: bytes
) -> tuple[list[ParsedRow], list[str], dict]:
    """
    An uploaded sheet or workbook, read the way the shop's catalog is
    actually organized: every product tab, in precedence order, duplicates
    resolved first-wins, retail packs merged.

    Returns (rows, warnings, report). The report is what the owner is shown
    after an upload — which tabs were read, which were skipped and why, how
    many rows each contributed, and every duplicate that was dropped.
    Uploading a sheet used to be an act of faith; this makes it an account.
    """
    sheets = _read_sheets(filename, content)
    retail = _read_retail_packs(filename, content)

    per_sheet: list[tuple[str, list[ParsedRow]]] = []
    warnings: list[str] = []
    sheet_stats: list[dict] = []

    for sheet_name, headers, data_rows in sheets:
        try:
            col_map = _find_column_map(headers)
        except CatalogParseError as exc:
            warnings.append(f"{sheet_name or 'sheet'}: {exc}")
            sheet_stats.append({"sheet": sheet_name, "rows": 0, "skipped": str(exc)})
            continue
        rows, sheet_warnings = _parse_rows(data_rows, col_map)
        per_sheet.append((sheet_name, rows))
        warnings.extend(f"{sheet_name}: {w}" if sheet_name else w for w in sheet_warnings)
        sheet_stats.append({"sheet": sheet_name, "rows": len(rows), "skipped": None})

    if not per_sheet:
        raise CatalogParseError(
            "None of the sheets in this file had readable price columns."
        )

    rows, dupe_warnings, clashes = dedupe_first_wins(per_sheet)
    warnings.extend(dupe_warnings)

    rows, attached, unmatched_retail = merge_retail_packs(rows, retail)

    report = {
        "sheets_read": sheet_stats,
        "sheets_skipped": sorted(SKIP_SHEETS),
        "duplicate_rows_dropped": len(clashes),
        "duplicate_details": clashes[:200],
        "full_bottle_prices_attached": attached,
        "full_bottle_unmatched": len(unmatched_retail),
        "full_bottle_unmatched_details": unmatched_retail,
        "products_after_merge": len(rows),
    }
    return rows, warnings, report


# --- ID + keyword generation ---------------------------------------------

def _slugify(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _make_unique_id(brand: str, name: str, used_ids: set[str]) -> str:
    base = _slugify(brand) + _slugify(name) or "perfume"
    candidate = base
    n = 2
    while candidate in used_ids:
        candidate = f"{base}_{n}"
        n += 1
    return candidate


def _ngrams(words: list[str], n: int) -> list[str]:
    return [" ".join(words[i : i + n]) for i in range(len(words) - n + 1)]


def _add_phrase_and_grams(keywords: set[str], text: str, corpus_stopwords: set[str]) -> None:
    norm = normalize_message(text)
    if not norm:
        return
    keywords.add(norm)
    words = norm.split()
    for n in (1, 2, 3):
        for gram in _ngrams(words, n):
            if len(gram) < 3:
                continue
            if n == 1 and (gram in GENERIC_STOPWORDS or gram in corpus_stopwords):
                continue
            keywords.add(gram)


def generate_keywords(brand: str, name: str, clone_of: str | None, corpus_stopwords: set[str]) -> list[str]:
    keywords: set[str] = set()
    _add_phrase_and_grams(keywords, f"{brand} {name}", corpus_stopwords)
    _add_phrase_and_grams(keywords, name, corpus_stopwords)
    if clone_of:
        _add_phrase_and_grams(keywords, clone_of, corpus_stopwords)
    return sorted(k for k in keywords if len(k) >= 3)


def _corpus_stopwords(parsed_rows: list[ParsedRow]) -> set[str]:
    """Words that show up in a large share of this upload's entries — too generic to be a safe standalone keyword."""
    freq: Counter = Counter()
    for r in parsed_rows:
        words = set(normalize_message(f"{r.brand} {r.name} {r.clone_of or ''}").split())
        freq.update(words)

    total = len(parsed_rows) or 1
    threshold = max(8, int(total * 0.015))
    return {w for w, c in freq.items() if c > threshold}


# --- Diff + candidate catalog construction --------------------------------

def build_catalog_from_rows(parsed_rows: list[ParsedRow], existing: dict[str, dict]) -> tuple[dict, dict]:
    """Build the candidate catalog dict + a structured diff against `existing`."""
    existing_by_name = {normalize_message(v.get("display_name", "")): pid for pid, v in existing.items()}
    corpus_stopwords = _corpus_stopwords(parsed_rows)

    new_catalog: dict[str, dict] = {}
    used_ids: set[str] = set()
    added: list[dict] = []
    updated: list[dict] = []

    for r in parsed_rows:
        display_name = f"{r.brand} {r.name}".strip()
        norm_name = normalize_message(display_name)
        existing_id = existing_by_name.get(norm_name)
        pid = existing_id or _make_unique_id(r.brand, r.name, used_ids)
        used_ids.add(pid)

        new_catalog[pid] = {
            "keywords": generate_keywords(r.brand, r.name, r.clone_of, corpus_stopwords),
            "display_name": display_name,
            "brand": r.brand or None,
            "prices": r.prices,
            "clone_of": r.clone_of,
        }

        if existing_id:
            old_prices = existing[existing_id].get("prices", {})
            if old_prices != r.prices:
                updated.append(
                    {
                        "perfume_id": pid,
                        "display_name": display_name,
                        "old_prices": old_prices,
                        "new_prices": r.prices,
                    }
                )
        else:
            added.append({"perfume_id": pid, "display_name": display_name, "prices": r.prices})

    removed = [
        {"perfume_id": pid, "display_name": v.get("display_name", pid)}
        for pid, v in existing.items()
        if pid not in used_ids
    ]

    diff = {
        "added": added,
        "updated": updated,
        "removed": removed,
        "added_count": len(added),
        "updated_count": len(updated),
        "removed_count": len(removed),
    }
    return new_catalog, diff


# --- Supabase-backed version storage ---------------------------------------

def _download_version_json(client, storage_path: str) -> dict:
    raw = client.storage.from_(CATALOG_BUCKET).download(storage_path)
    return json.loads(raw.decode("utf-8"))


def _get_active_catalog(client) -> dict:
    """The catalog to diff a new upload against: Supabase's active version if one exists, else the live in-memory catalog."""
    resp = client.table("catalog_versions").select("storage_path").eq("is_active", True).limit(1).execute()
    if resp.data:
        return _download_version_json(client, resp.data[0]["storage_path"])

    from app.catalog import PERFUMES

    return dict(PERFUMES)


def _get_version(client, version_id: int) -> dict | None:
    resp = client.table("catalog_versions").select("*").eq("id", version_id).limit(1).execute()
    return resp.data[0] if resp.data else None


def _write_catalog_file(catalog: dict) -> None:
    from app.catalog import CATALOG_PATH

    CATALOG_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")


def create_pending_version(filename: str, content: bytes) -> dict:
    """Parse an uploaded sheet, diff it against the active catalog, and store it as a pending version for review."""
    client = require_client()

    active = _get_active_catalog(client)
    parsed_rows, row_warnings, report = parse_workbook(filename, content)

    if not parsed_rows:
        raise CatalogParseError("No usable rows found in the uploaded sheet.")

    new_catalog, diff = build_catalog_from_rows(parsed_rows, active)
    diff["sheet_report"] = report

    insert_resp = (
        client.table("catalog_versions")
        .insert(
            {
                "status": "pending",
                "source_filename": filename,
                "storage_path": "",
                "perfume_count": len(new_catalog),
                "added_count": diff["added_count"],
                "updated_count": diff["updated_count"],
                "removed_count": diff["removed_count"],
                "diff": diff,
                "parse_warnings": row_warnings,
            }
        )
        .execute()
    )
    version = insert_resp.data[0]
    version_id = version["id"]
    storage_path = f"v{version_id}.json"

    blob = json.dumps(new_catalog, ensure_ascii=False, indent=2).encode("utf-8")
    client.storage.from_(CATALOG_BUCKET).upload(
        storage_path, blob, {"content-type": "application/json", "upsert": "true"}
    )
    client.table("catalog_versions").update({"storage_path": storage_path}).eq("id", version_id).execute()
    version["storage_path"] = storage_path

    return version


def list_versions(limit: int = 30) -> list[dict]:
    client = require_client()
    resp = (
        client.table("catalog_versions")
        .select(
            "id,status,is_active,source_filename,perfume_count,added_count,"
            "updated_count,removed_count,parse_warnings,created_at,published_at"
        )
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    return resp.data or []


def get_version_detail(version_id: int) -> dict:
    client = require_client()
    version = _get_version(client, version_id)
    if version is None:
        raise CatalogParseError(f"Version {version_id} not found")
    return version


def _activate_version(client, version_id: int) -> dict:
    version = _get_version(client, version_id)
    if version is None:
        raise CatalogParseError(f"Version {version_id} not found")

    catalog = _download_version_json(client, version["storage_path"])
    _write_catalog_file(catalog)

    from app.catalog import reload_catalog

    reload_catalog()

    now = datetime.now(timezone.utc).isoformat()
    client.table("catalog_versions").update({"is_active": False}).eq("is_active", True).execute()
    client.table("catalog_versions").update(
        {"status": "published", "is_active": True, "published_at": now}
    ).eq("id", version_id).execute()

    version["status"] = "published"
    version["is_active"] = True
    version["published_at"] = now
    return version


# A publish that deletes more than this many products is treated as an
# accident until the owner says otherwise. Not a guess at a safe number —
# a real upload of the shop's own workbook removed 208 products because the
# parser read one sheet of six, and the only thing standing between that
# and a live catalog was a diff screen where "removed: 208" looked like a
# statistic. Routine edits remove nothing; a discontinued line removes a
# handful. Anything larger deserves a sentence, not a click.
MAX_SILENT_REMOVALS = 10


def publish_version(version_id: int, confirm_removals: bool = False) -> dict:
    """
    Make a pending version live: writes catalog_data.json and hot-reloads
    the running bot.

    Refuses outright when the version would delete a large part of the
    catalog, unless the caller confirms it. The confirmation exists so the
    owner has to have read the number.
    """
    client = require_client()
    version = _get_version(client, version_id)
    if version is None:
        raise CatalogParseError(f"Version {version_id} not found")
    if version["status"] != "pending":
        raise CatalogParseError(f"Version {version_id} is '{version['status']}', not pending — nothing to publish")

    removed = version.get("removed_count") or 0
    if removed > MAX_SILENT_REMOVALS and not confirm_removals:
        raise CatalogRemovalWarning(
            f"This would remove {removed} products from the live catalog. That is "
            f"usually a sign the upload is missing a sheet rather than that the "
            f"shop stopped selling {removed} perfumes — check the removed list "
            f"before confirming.",
            removed=removed,
        )

    return _activate_version(client, version_id)


def rollback_version(version_id: int) -> dict:
    """Re-activate a previously-published version (any version, active or not)."""
    client = require_client()
    version = _get_version(client, version_id)
    if version is None:
        raise CatalogParseError(f"Version {version_id} not found")
    if version["status"] != "published":
        raise CatalogParseError(f"Version {version_id} was never published — nothing to roll back to")
    return _activate_version(client, version_id)


def discard_version(version_id: int) -> None:
    """Reject a pending version without ever making it live."""
    client = require_client()
    client.table("catalog_versions").update({"status": "discarded"}).eq("id", version_id).eq(
        "status", "pending"
    ).execute()


def sync_active_catalog_to_disk() -> bool:
    """
    Best-effort startup hook: pull whatever is active in Supabase down to
    catalog_data.json and hot-load it, so a redeploy picks up the latest
    published catalog instead of whatever was baked into the deploy image.

    No-ops (returns False) if Supabase isn't configured or has no active
    version yet — the bundled catalog_data.json keeps working either way.
    """
    try:
        client = require_client()
    except Exception:
        return False

    resp = client.table("catalog_versions").select("storage_path").eq("is_active", True).limit(1).execute()
    if not resp.data:
        return False

    catalog = _download_version_json(client, resp.data[0]["storage_path"])
    _write_catalog_file(catalog)

    from app.catalog import reload_catalog

    reload_catalog()
    return True
